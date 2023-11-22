import os
import openpyxl
import pandas as pd
from datetime import datetime
from copy import copy
import webbrowser

def get_filenames_in(path, file_extension=None):
    filenames = os.listdir(path)
    if file_extension:
        filenames = [filename for filename in filenames if filename.endswith(file_extension) and not filename.startswith("seperated_")]
    return filenames

def convert2dataframe(wb):
    ws = wb.worksheets[0]
    data = ws.values
    cols = next(data) # 첫 번째 행을 cols로 생각
    data = [d for d in data]
    return pd.DataFrame(data, columns=cols)

def get_date_from_timestamp(timestamp):
    return datetime.strptime(str(timestamp), "%Y-%m-%d %H:%M:%S").date()

def replace_row_with(values, ws, rowidx, first_cells):
    print(f"\rWriting {rowidx} {values[0]}", end="")
    for idx, value in enumerate(values):
        ws.cell(row=rowidx, column=idx+1).font = copy(first_cells[idx].font)
        ws.cell(row=rowidx, column=idx+1).border = copy(first_cells[idx].border)
        ws.cell(row=rowidx, column=idx+1).fill = copy(first_cells[idx].fill)
        ws.cell(row=rowidx, column=idx+1).number_format = copy(first_cells[idx].number_format)
        ws.cell(row=rowidx, column=idx+1).protection = copy(first_cells[idx].protection)
        ws.cell(row=rowidx, column=idx+1).alignment = copy(first_cells[idx].alignment)
        ws.cell(row=rowidx, column=idx+1).value = value

def save_dataframe_with_space(wb, df, filename, space):
    ws = wb.worksheets[0]
    
    rowidx = 2
    predate = get_date_from_timestamp(df.iloc[0, 0])
    first_cells = [ws.cell(row=2, column=1+i) for i in range(0, len(df.columns))] # 셀서식 유지를 위해 복사!
    for _, row in df.iterrows():
        curdate = get_date_from_timestamp(row.iloc[0])
        if curdate != predate:
            predate = curdate
            for i in range(0, space):
                replace_row_with([None]*len(row), ws, rowidx, first_cells)
                rowidx += 1
        replace_row_with(row.tolist(), ws, rowidx, first_cells)
        rowidx += 1

    wb.save(filename=f"seperated_{filename}")

def main():
    filenames = get_filenames_in(path="./", file_extension="xlsx")

    print("------------ 변환할 xlsx 파일 목록 -----------")
    for filename in filenames:
        print(f"- {filename}")
    print("----------------------------------------------")
    space = input("공백라인 수 입력후 Enter를 쳐주세요 (미입력시 3개) : ")
    if space == "":
        space = 3
    else:
        space = int(space)

    print(f"공백라인 수 {space} 개로 작업 시작합니다!")

    for filename in filenames:
        print(f"\r{filename} 처리중...")
        wb = openpyxl.load_workbook(filename=filename)
        df = convert2dataframe(wb)
        save_dataframe_with_space(wb, df, filename, space=space)
        wb.close()
        print(f"\r{filename} 완료                              ")

    webbrowser.open("https://coding-hwangsawon.tistory.com/3")

try:
    main()
except Exception as err:
    print("")
    print("에러발생")
    print(f"Unexpected {err}, {type(err)}")
    print("")

# To protect exit()
input()